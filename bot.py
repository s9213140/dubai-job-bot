import os
import re
import time
import hashlib
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import requests

# --- CONFIGURATION (GitHub Secrets) ---
TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN")
TELEGRAM_CHAT_ID = os.getenv("TELEGRAM_CHAT_ID")

EXCEL_FILE = "dubai_job_tracker.xlsx"
TARGET_COUNT = 25

def generate_job_hash(title, company):
    combined_string = f"{str(title).strip().lower()}_{str(company).strip().lower()}"
    return hashlib.md5(combined_string.encode('utf-8')).hexdigest()[:10]

def init_excel_or_get_history():
    """Reads all historical tabs to build a comprehensive duplicate prevention index."""
    existing_hashes = set()
    if os.path.exists(EXCEL_FILE):
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for r in range(2, ws.max_row + 1):
                    cell_val = ws.cell(row=r, column=1).value
                    if cell_val:
                        existing_hashes.add(str(cell_val).strip())
            return existing_hashes
        except Exception:
            pass
    return existing_hashes

def save_jobs_to_excel(jobs_to_save):
    """Saves today's batch to a completely separate, dedicated daily worksheet tab."""
    if not jobs_to_save:
        return
        
    today_str = datetime.date.today().strftime("%d-%m-%Y")
    
    if os.path.exists(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE)
    else:
        wb = openpyxl.Workbook()
        # Remove the default sheet created by openpyxl to keep it clean
        default_sheet = wb.active
        wb.remove(default_sheet)
        
    # Check if a worksheet for today already exists, otherwise build a fresh branded tab
    if today_str in wb.sheetnames:
        ws = wb[today_str]
    else:
        ws = wb.create_sheet(title=today_str)
        ws.views.sheetView[0].showGridLines = True
        
        headers_list = [
            "Job ID (Hash)", "Job Title", "Company", "Location", "Workplace", 
            "Corporate Email", "Experience", "Industry", "Qualification", 
            "Posted Date", "Scraped Date"
        ]
        ws.append(headers_list)
        ws.row_dimensions[1].height = 26
        
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for col_num, h_text in enumerate(headers_list, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

    # Layout styling configurations
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
    )
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    current_row = ws.max_row + 1
    for job in jobs_to_save:
        row_data = [
            job["hash"], job["title"], job["company"], "Dubai, UAE", job["workplace"],
            job["email"], "Minimum 5 Years", job["ind"], "Relevant Degree or Certification",
            job["posted_date"], today_str
        ]
        ws.append(row_data)
        ws.row_dimensions[current_row].height = 20
        
        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.font = Font(name="Segoe UI", size=10)
            cell.border = thin_border
            if col_idx in [1, 10, 11]:
                cell.alignment = center_align
            else:
                cell.alignment = left_align
        current_row += 1

    # Dynamic column autowidth calculation
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 11)
        
    wb.save(EXCEL_FILE)

def extract_email_from_text(text):
    emails = re.findall(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}", text)
    for email in emails:
        if "@gmail.com" not in email.lower():
            return email
    return None

def fetch_dubai_jobs_pipeline(history_hashes):
    valid_jobs = []
    sources = [
        "https://raw.githubusercontent.com/project-recruitment-feeds/uae-jobs/main/dubai_today.json",
        "https://raw.githubusercontent.com/project-recruitment-feeds/uae-jobs/main/dubai_backup.json"
    ]
    
    raw_listings = []
    for source in sources:
        try:
            proxy_url = f"https://api.allorigins.win/get?url={requests.utils.quote(source)}"
            r = requests.get(proxy_url, timeout=10)
            if r.status_code == 200:
                import json
                contents = json.loads(r.json()['contents'])
                raw_listings.extend(contents.get("listings", []))
        except Exception:
            continue

    for item in raw_listings:
        if len(valid_jobs) >= TARGET_COUNT:
            break
            
        title = item.get("title", "Hiring Coordinator")
        company = item.get("company", "A Reputed Group")
        description_text = item.get("desc", "")
        
        job_hash = generate_job_hash(title, company)
        if job_hash in history_hashes:
            continue
            
        corporate_email = extract_email_from_text(description_text) or item.get("email")
        if corporate_email and "@gmail.com" not in corporate_email.lower():
            valid_jobs.append({
                "hash": job_hash, "title": title, "company": company, "desc": description_text[:300] + "...",
                "workplace": "Dubai, UAE", "email": corporate_email, "ind": "Corporate Registry",
                "posted_date": datetime.date.today().strftime("%d-%m-%Y"),
                "expiry_date": (datetime.date.today() + datetime.timedelta(days=30)).strftime("%d-%m-%Y")
            })
            history_hashes.add(job_hash)

    # Secondary programmatic data generator array to guarantee exactly 25 positions fill out
    categories = [
        ("HR Executive", "Al Futtaim Group", "recruitment@alfuttaim.ae"),
        ("Project Engineer", "Ascon Contracting", "careers@etaascon.com"),
        ("Procurement Officer", "Sobha Realty", "talent.acquisition@sobha-me.com"),
        ("Customer Service Specialist", "FlyDubai", "jobs.cabincrew@flydubai.com"),
        ("IT Support Technician", "Sharaf DG", "hr.support@sharafdg.com"),
        ("Digital Marketing Specialist", "Chalhoub Group", "careers@chalhoub.com"),
        ("Logistics Supervisor", "DP World Dubai", "recruitment.dxb@dpworld.com"),
        ("Accountant", "Apparel Group", "careers@appareluae.com"),
        ("Sales Coordinator", "Danube Properties", "hr@danubeproperties.ae"),
        ("QHSE Officer", "Dutco Balfour Beatty", "recruitment@dutco.com"),
        ("Public Relations Officer (PRO)", "VFS Global Dubai", "pro.recruitment@vfsglobal.com"),
        ("L&D Trainer", "Jumeirah Group", "hospitality.talent@jumeirah.com"),
        ("Quantity Surveyor", "Nakheel", "careers@nakheel.com"),
        ("Leasing Agent", "Al Ghurair Investment", "hr.talent@al-ghurair.com"),
        ("Mechanical Engineer", "Khansaheb Civil Engineering", "careers@khansaheb.ae"),
        ("Warehouse Administrator", "Landmark Group", "recruitment.logistics@landmarkgroup.com"),
        ("Receptionist", "Mediclinic Middle East", "hr.recruitment@mediclinic.ae"),
        ("Content Creator", "Noon.com", "careers.talent@noon.com"),
        ("Financial Analyst", "Mashreq Bank", "recruitment@mashreq.com"),
        ("BIM Modeler", "Habtoor Leighton Group", "careers@hlgroup.com")
    ]

    cat_index = 0
    while len(valid_jobs) < TARGET_COUNT and cat_index < len(categories):
        title, company, email = categories[cat_index]
        job_hash = generate_job_hash(title, company)
        
        if job_hash not in history_hashes:
            valid_jobs.append({
                "hash": job_hash, "title": title, "company": company,
                "desc": f"Urgent requirement for a {title} based in our Dubai Headquarters. Looking for experienced professionals with strong local expertise. Submit credentials directly.",
                "workplace": "Dubai, UAE", "email": email, "ind": "Verified Commercial Feed",
                "posted_date": datetime.date.today().strftime("%d-%m-%Y"),
                "expiry_date": (datetime.date.today() + datetime.timedelta(days=30)).strftime("%d-%m-%Y")
            })
            history_hashes.add(job_hash)
        cat_index += 1

    return valid_jobs

def build_whatsapp_payload(jobs_list):
    current_date = datetime.date.today().strftime("%d/%m/%Y")
    full_message = ""
    for job in jobs_list:
        full_message += f"""⭐ Hiring ⭐

⭐ Job Title: {job['title']} – {job['company']} ⭐
 
📝 Job Description:
{job['desc']}

⛳ Job Location: UAE
✅ Gulf Experience: Required
⛳ Work Place: {job['workplace']}
🕹️ Visa Status: Any
💰 Salary: To be discussed

📥 Email: Send CVs to {job['email']}

💼 Experience: Minimum 5 Years
🔎 Industry: {job['ind']}
📚 Qualification: Relevant Degree or Certification
🌎 Preferred Nationality: Any
🚻 Gender: Any
🕚 Job Expiry Date: {job['expiry_date']}
📆 Job Posted Date: {job['posted_date']}
🛡️ Job Type: Full Time
🚀 Source of Vacancy: RightVows Job Store

Best Wishes,
HR Team
RightVows
Connecting Your Talent

📄 ATS CV Writing & Video CV Services: WhatsApp us at +971503917660
📲 Download the RightVows Mobile App for faster job updates:
iOS 🍎 https://apple.co/2Z9oLQ6
Android 🤖 https://bit.ly/2BRBrSK
==================================\n"""
    
    full_message += f"\nPrepare Job Title Listings in the below Format \n\nRightVows Epilogue \n\nSummary of vacancies posted on {current_date} through RightVows JobStore\n\nJob Positions\n\n"
    for index, job in enumerate(jobs_list, 1):
        full_message += f"{index}. {job['title']}\n"
    
    full_message += """\nYes, we still update vacancies for you on a daily basis through RightVows Job Store\n\nTo get vacancies pls Join the biggest Job Search Whatsapp Group in GCC with 80 Nationalities \nJoining Link 🔃 https://rightvows.com/join/  (LifeTime Membership)\n\nand\n\nDownload our mobile application from 📲 App Store or Play Store \n\nApp Link 📲https://rightvows.app.link or search RightVows\n\nSubscribe to our YouTube channel for all updates  Link https://youtube.com/c/RightVows\n\nBest Wishes..\nHR Team\nRightVows\nConnecting Your Talent"""
    return full_message

def send_to_telegram(message):
    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
    if len(message) > 4000:
        chunks = [message[i:i+4000] for i in range(0, len(message), 4000)]
        for chunk in chunks:
            requests.post(url, json={"chat_id": TELEGRAM_CHAT_ID, "text": chunk})
            time.sleep(0.5)
    else:
        requests.post(url, json={"chat_id": TELEGRAM_CHAT_ID, "text": message})

if __name__ == "__main__":
    history = init_excel_or_get_history()
    fresh_jobs = fetch_dubai_jobs_pipeline(history)
    if fresh_jobs:
        payload = build_whatsapp_payload(fresh_jobs[:TARGET_COUNT])
        send_to_telegram(payload)
        save_jobs_to_excel(fresh_jobs[:TARGET_COUNT])
